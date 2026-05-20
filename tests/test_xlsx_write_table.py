"""Unit tests for write_table claim_id comment propagation."""

from pathlib import Path

import pandas as pd
import pytest
from openpyxl import Workbook, load_workbook
from openpyxl.comments import Comment

from scripts.xlsx import write_table, new_close_pack
from scripts.xlsx.qa import assert_full_claim_coverage, claim_id_coverage


def _make_pack(tmp_dir: Path) -> Path:
    pack = tmp_dir / "test_pack.xlsx"
    new_close_pack(pack, "2026-05")
    return pack


# ---------------------------------------------------------------------------
# claim_id_per_row: basic coverage
# ---------------------------------------------------------------------------

def test_write_table_numeric_cells_get_claim_comments_per_row(tmp_path):
    """Every numeric cell in a data row must carry a claim_id comment when
    claim_id_per_row is supplied."""
    pack = _make_pack(tmp_path)
    df = pd.DataFrame({
        "label": ["Revenue", "Opex", "Net Income"],
        "amount_usd": [1_000_000.0, -600_000.0, 400_000.0],
    })
    claim_ids = [
        "controller.pnl.line.revenue",
        "controller.pnl.line.opex",
        "controller.pnl.line.net_income",
    ]
    write_table(pack, "P&L Test", df, claim_id_per_row=claim_ids)

    wb = load_workbook(pack, data_only=False)
    ws = wb["P&L Test"]

    # Row 1 is the header — no comment expected on any cell.
    for cell in ws[1]:
        assert cell.comment is None, f"Header cell {cell.coordinate} should have no comment"

    # Data rows 2..4 — the numeric column (column 2: amount_usd) must have a
    # claim_id comment; the text column (column 1: label) must not.
    for row_offset, expected_id in enumerate(claim_ids):
        r = row_offset + 2  # row 2 = first data row
        text_cell = ws.cell(row=r, column=1)
        numeric_cell = ws.cell(row=r, column=2)
        assert text_cell.comment is None, \
            f"Text cell {text_cell.coordinate} should not receive a comment"
        assert numeric_cell.comment is not None, \
            f"Numeric cell {numeric_cell.coordinate} missing claim_id comment"
        assert f"claim_id: {expected_id}" in numeric_cell.comment.text, \
            f"Wrong claim_id on {numeric_cell.coordinate}: {numeric_cell.comment.text!r}"


def test_write_table_multi_numeric_columns_all_tagged(tmp_path):
    """When a row has multiple numeric columns, all of them get the row's
    claim_id comment."""
    pack = _make_pack(tmp_path)
    df = pd.DataFrame({
        "account": ["4100", "6300"],
        "actual_usd": [2_556_500.0, -332_400.0],
        "budget_usd": [2_951_000.0, -272_400.0],
        "variance_usd": [-394_500.0, -60_000.0],
        "variance_pct": [-0.1337, 0.2203],
    })
    claim_ids = [
        "fpa.variance_vs_budget.4100.usd",
        "fpa.variance_vs_budget.6300.usd",
    ]
    write_table(pack, "Var Test", df, claim_id_per_row=claim_ids)

    wb = load_workbook(pack, data_only=False)
    ws = wb["Var Test"]

    numeric_cols = [2, 3, 4, 5]  # actual_usd, budget_usd, variance_usd, variance_pct
    for row_offset, cid in enumerate(claim_ids):
        r = row_offset + 2
        for col in numeric_cols:
            cell = ws.cell(row=r, column=col)
            assert cell.comment is not None, \
                f"Numeric cell {cell.coordinate} missing claim_id comment"
            assert f"claim_id: {cid}" in cell.comment.text


def test_write_table_passes_full_coverage_gate(tmp_path):
    """A workbook built with claim_id_per_row must satisfy assert_full_claim_coverage."""
    pack = _make_pack(tmp_path)
    df = pd.DataFrame({
        "pnl_line": ["Revenue / Subscription", "Opex / Salaries"],
        "amount_usd": [2_556_500.0, -1_362_000.0],
    })
    write_table(pack, "P&L", df, claim_id_per_row=[
        "controller.pnl.line.revenue_subscription",
        "controller.pnl.line.opex_salaries",
    ])
    # Must not raise.
    assert_full_claim_coverage(pack)


# ---------------------------------------------------------------------------
# Backward-compatibility: callers that omit claim_id_per_row still work
# ---------------------------------------------------------------------------

def test_write_table_without_claim_ids_still_writes(tmp_path):
    """Omitting both claim arguments must not raise — legacy callers unaffected."""
    pack = _make_pack(tmp_path)
    df = pd.DataFrame({"x": [1.0, 2.0], "y": ["a", "b"]})
    write_table(pack, "Legacy Sheet", df)  # no claim args
    wb = load_workbook(pack, data_only=False)
    assert "Legacy Sheet" in wb.sheetnames


def test_write_table_without_claim_ids_produces_orphans(tmp_path):
    """Omitting claim args does produce orphan cells — so callers MUST supply
    claim_id_per_row on production paths or the gate will reject the pack."""
    pack = _make_pack(tmp_path)
    df = pd.DataFrame({"amount_usd": [1_000.0, 2_000.0]})
    write_table(pack, "Orphan Sheet", df)
    cov = claim_id_coverage(pack)
    assert len(cov["orphans"]) > 0, \
        "Numeric cells without claim_ids must appear as orphans in coverage report"


# ---------------------------------------------------------------------------
# claim_ids_by_cell: legacy explicit-cell mode still works
# ---------------------------------------------------------------------------

def test_write_table_claim_ids_by_cell_still_works(tmp_path):
    """The legacy claim_ids_by_cell dict interface must still attach comments."""
    pack = _make_pack(tmp_path)
    df = pd.DataFrame({"amount": [42.0]})
    write_table(pack, "Cell Mode", df, claim_ids_by_cell={"B2": "my.claim.id"})
    wb = load_workbook(pack, data_only=False)
    ws = wb["Cell Mode"]
    assert ws["B2"].comment is not None
    assert "claim_id: my.claim.id" in ws["B2"].comment.text
