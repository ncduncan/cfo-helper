"""
Tests for per-column numeric validation in connectors/excel.py.

Stray spaces, currency symbols, or text in numeric columns must raise
ValueError naming the workbook, the column, and a sample of the bad cells.
Clean numeric data must pass through and arrive as float dtype.

Notes on test fixture design
-----------------------------
When all values in an Excel column are numeric-looking strings (e.g. "1234.5 "
with a trailing space), pandas/openpyxl auto-promotes the column to float64 on
read, silently dropping the space. The stray-space scenario therefore requires a
mixed-type column (at least one genuinely non-numeric value alongside the
space-tainted value) to keep pandas from coercing the whole column to float.
In practice the bug manifests the same way; the test fixture just needs to
replicate the mixed-column case that arises from partial CSV imports.

The validate_numeric tests (unit-level) exercise _validate_numeric directly
with a manually crafted DataFrame so they are independent of the read path.
"""

from pathlib import Path
import pandas as pd
import pytest
from openpyxl import Workbook
from connectors import excel


def _inputs_dir(tmp_path: Path) -> Path:
    """Create and return the workspace/inputs subdirectory _resolve_workbook expects."""
    d = tmp_path / "inputs"
    d.mkdir(exist_ok=True)
    return d


# ---------------------------------------------------------------------------
# Unit tests: _validate_numeric in isolation
# ---------------------------------------------------------------------------

def test_validate_numeric_rejects_stray_space():
    """_validate_numeric must raise when a string with trailing whitespace
    is present in a declared-numeric column."""
    df = pd.DataFrame({
        "entity": ["UK"],
        "period": ["2026-05"],
        "account": ["4100"],
        "account_name": ["Sub Rev"],
        "debit": [0.0],
        "credit": [1000.0],
        "currency": ["GBP"],
        "amount_local": [1000.0],
        "amount_usd": ["1234.5 "],   # stray trailing space, kept as object
    })
    fake_wb = Path("/fake/workspace/inputs/GL_bad.xlsx")
    with pytest.raises(ValueError) as exc:
        excel._validate_numeric(df, "gl", fake_wb)
    msg = str(exc.value)
    assert "GL_bad.xlsx" in msg
    assert "amount_usd" in msg
    assert "1234.5 " in msg or "non-numeric" in msg.lower()


def test_validate_numeric_clean_passes_through():
    """_validate_numeric must return a DataFrame with float dtype for clean data."""
    df = pd.DataFrame({
        "entity": ["UK"],
        "period": ["2026-05"],
        "account": ["4100"],
        "account_name": ["Sub Rev"],
        "debit": [0.0],
        "credit": [1000.0],
        "currency": ["GBP"],
        "amount_local": [1000.0],
        "amount_usd": [1234.5],
    })
    fake_wb = Path("/fake/workspace/inputs/GL_ok.xlsx")
    out = excel._validate_numeric(df, "gl", fake_wb)
    assert out["amount_usd"].dtype.kind == "f"
    assert out.loc[0, "amount_usd"] == 1234.5


# ---------------------------------------------------------------------------
# Integration tests: _read_typed end-to-end through Excel read path
# ---------------------------------------------------------------------------

def test_clean_numeric_passes_through(tmp_path):
    """_read_typed must return float dtype for a clean GL workbook."""
    inputs = _inputs_dir(tmp_path)
    wb = Workbook()
    ws = wb.active
    ws.title = "TB"
    ws.append([
        "entity", "period", "account", "account_name", "debit", "credit",
        "currency", "amount_local", "amount_usd",
    ])
    ws.append(["UK", "2026-05", "4100", "Subscription Revenue",
               0, 1000, "GBP", 1000, 1234.5])
    p = inputs / "GL_ok.xlsx"
    wb.save(p)
    spec = {"workbook": "GL_ok.xlsx", "sheet": "TB", "header_row": 1}
    df = excel._read_typed(tmp_path, spec, list(excel.GL_COLUMNS), "gl")
    assert df["amount_usd"].dtype.kind == "f"  # numeric float
    assert df.loc[0, "amount_usd"] == 1234.5


def test_currency_symbol_in_numeric_raises(tmp_path):
    """A dollar-sign prefix is also a non-numeric value and must be caught."""
    inputs = _inputs_dir(tmp_path)
    wb = Workbook()
    ws = wb.active
    ws.title = "Closed"
    ws.append(["deal_id", "customer_id", "customer_name", "period", "stage",
               "tcv_usd", "acv_usd", "product", "owner"])
    ws.append(["D001", "C001", "Acme", "2026-05", "closed_won",
               "$500000", 50000, "Flight Ops", "rep1"])
    p = inputs / "Deals_bad.xlsx"
    wb.save(p)
    spec = {"workbook": "Deals_bad.xlsx", "sheet": "Closed", "header_row": 1}
    with pytest.raises(ValueError) as exc:
        excel._read_typed(tmp_path, spec, list(excel.DEAL_COLUMNS), "deals")
    msg = str(exc.value)
    assert "Deals_bad.xlsx" in msg
    assert "tcv_usd" in msg


def test_assumption_workbook_rejects_dirty_period_amount(tmp_path):
    """A dirty cell in `period_amount_usd` (the assumption schema's numeric
    column, per connectors/assumptions.py) must raise. This test guards the
    domain key in NUMERIC_COLUMNS — earlier code keyed on the wrong column
    name, silently skipping validation."""
    from connectors.assumptions import ASSUMPTION_COLUMNS
    inputs = _inputs_dir(tmp_path)
    wb = Workbook()
    ws = wb.active
    ws.title = "Plan"
    ws.append(list(ASSUMPTION_COLUMNS))
    ws.append([
        "UK", "2026-05", "plan_fy26", "4100",
        "Subscription Revenue", "revenue",
        "flight_ops", "sales",
        "tail_count", 100,
        10, 100, "$10K",   # bad period_amount_usd
        "2026-05-01", "plan.xlsx",
    ])
    p = inputs / "Plan_bad.xlsx"
    wb.save(p)
    spec = {"workbook": "Plan_bad.xlsx", "sheet": "Plan", "header_row": 1}
    with pytest.raises(ValueError) as exc:
        excel._read_typed(tmp_path, spec, list(ASSUMPTION_COLUMNS), "assumptions")
    msg = str(exc.value)
    assert "Plan_bad.xlsx" in msg
    assert "period_amount_usd" in msg
    assert "$10K" in msg or "non-numeric" in msg.lower()


def test_multiple_bad_rows_sample_appears_in_error(tmp_path):
    """Error message must include a sample of the offending values."""
    inputs = _inputs_dir(tmp_path)
    wb = Workbook()
    ws = wb.active
    ws.title = "Rates"
    ws.append(["currency", "period", "rate_to_usd_avg", "rate_to_usd_eop"])
    ws.append(["GBP", "2026-05", "N/A", "1.27"])
    ws.append(["EUR", "2026-05", "TBD", "1.09"])
    p = inputs / "FX_bad.xlsx"
    wb.save(p)
    spec = {"workbook": "FX_bad.xlsx", "sheet": "Rates", "header_row": 1}
    with pytest.raises(ValueError) as exc:
        excel._read_typed(tmp_path, spec, list(excel.FX_COLUMNS), "fx")
    msg = str(exc.value)
    assert "FX_bad.xlsx" in msg
    assert "rate_to_usd_avg" in msg
    # At least one of the bad values should appear in the error
    assert "N/A" in msg or "TBD" in msg
