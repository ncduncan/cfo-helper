"""
QA checks for cfo-helper Excel output.

# Pattern source: anthropics/skills/document-skills/xlsx — their skill
# enforces "no formula errors, every numeric cell tied to a source." We
# adapt that to "every numeric cell carries a claim_id comment" (our
# project-specific provenance contract per CLAUDE.md §8 rule 2).
"""

from __future__ import annotations

from pathlib import Path
from typing import Iterator

from openpyxl import load_workbook
from openpyxl.cell.cell import Cell

from scripts.xlsx.recalc import iter_error_cells


def _is_numeric(value) -> bool:
    if isinstance(value, bool):
        return False
    return isinstance(value, (int, float))


def iter_orphan_numeric_cells(path: Path) -> Iterator[tuple[str, str, object]]:
    """Yield (sheet, cell_ref, value) for every numeric cell missing a claim_id comment.

    Cells whose value is a formula are skipped (formulas derive from
    claim-tagged inputs; their provenance lives on the upstream cell).
    """
    wb = load_workbook(path, data_only=False)
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        for row in ws.iter_rows():
            for cell in row:
                if not isinstance(cell, Cell):
                    continue
                # Skip header band by convention (row 1 of sheets that have headers)
                if cell.row == 1:
                    continue
                # Skip formula cells
                if isinstance(cell.value, str) and cell.value.startswith("="):
                    continue
                if not _is_numeric(cell.value):
                    continue
                if cell.comment is None or "claim_id:" not in (cell.comment.text or ""):
                    yield (sheet_name, cell.coordinate, cell.value)


def claim_id_coverage(path: Path) -> dict:
    """Summarize claim-id coverage across a workbook.

    Returns a dict with: total_numeric, with_claim_id, orphans (list of
    (sheet, ref, value)). Use as a self-check input.
    """
    wb = load_workbook(path, data_only=False)
    total = 0
    tagged = 0
    orphans: list[tuple[str, str, object]] = []
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        for row in ws.iter_rows():
            for cell in row:
                if cell.row == 1:
                    continue
                if isinstance(cell.value, str) and cell.value.startswith("="):
                    continue
                if not _is_numeric(cell.value):
                    continue
                total += 1
                if cell.comment and "claim_id:" in (cell.comment.text or ""):
                    tagged += 1
                else:
                    orphans.append((sheet_name, cell.coordinate, cell.value))
    return {
        "total_numeric": total,
        "with_claim_id": tagged,
        "coverage_pct": (tagged / total) if total else 1.0,
        "orphans": orphans,
    }


def assert_full_claim_coverage(path: Path) -> None:
    """Raise AssertionError if any numeric input cell lacks a claim_id comment."""
    summary = claim_id_coverage(path)
    if summary["orphans"]:
        formatted = "\n".join(
            f"  {s}!{c}: {v}" for s, c, v in summary["orphans"][:20]
        )
        more = "" if len(summary["orphans"]) <= 20 else f"\n  ...and {len(summary['orphans']) - 20} more"
        raise AssertionError(
            f"Workbook {path} has {len(summary['orphans'])} numeric cell(s) "
            f"without claim_id comments:\n{formatted}{more}"
        )


__all__ = [
    "iter_orphan_numeric_cells",
    "claim_id_coverage",
    "assert_full_claim_coverage",
    "iter_error_cells",  # convenience re-export
]
