"""
LibreOffice-based formula recalc + error scan for cfo-helper close packs.

# Pattern source: anthropics/skills/document-skills/xlsx/scripts/recalc.py —
# their tool uses headless LibreOffice to force formula recalculation and
# detect cell errors. We re-implement here so it's optional: if `soffice`
# isn't on PATH (locked-down corporate environment may lack LibreOffice),
# we log a warning and skip recalc but still scan for errors via openpyxl.
"""

from __future__ import annotations

import logging
import shutil
import subprocess
from pathlib import Path
from typing import Iterator

from openpyxl import load_workbook

logger = logging.getLogger(__name__)

ERROR_VALUES = {"#REF!", "#DIV/0!", "#VALUE!", "#NAME?", "#NUM!", "#N/A", "#NULL!"}


def _soffice_path() -> str | None:
    return shutil.which("soffice") or shutil.which("libreoffice")


def recalc_workbook(path: Path, *, timeout: int = 90) -> bool:
    """Force-recalc all formulas in `path` via headless LibreOffice.

    Returns True if recalc ran, False if soffice was not found (caller can
    decide whether to treat this as fatal). Errors during recalc raise.
    """
    soffice = _soffice_path()
    if not soffice:
        logger.warning(
            "LibreOffice (soffice) not found on PATH; skipping formula recalc for %s. "
            "Install LibreOffice or run the workbook through Excel before sharing.",
            path,
        )
        return False

    macro = (
        "macro:///Standard.Module1.RecalcAndSave"  # may not exist; we use --headless
    )
    # Simpler: convert in place with --calc + --norestore, which forces a recalc.
    cmd = [
        soffice,
        "--headless",
        "--calc",
        "--norestore",
        "--nologo",
        "--nofirststartwizard",
        "--convert-to", "xlsx",
        "--outdir", str(path.parent),
        str(path),
    ]
    subprocess.run(cmd, check=True, timeout=timeout, capture_output=True)
    return True


def iter_error_cells(path: Path) -> Iterator[tuple[str, str, str]]:
    """Yield (sheet, cell_ref, error_value) for every cell holding an Excel error.

    Loads with `data_only=True` so we read cached calculated values, not
    formula strings. If the workbook has never been opened in a calc engine,
    cached values may be missing — run `recalc_workbook` first.
    """
    wb = load_workbook(path, data_only=True)
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        for row in ws.iter_rows():
            for cell in row:
                if cell.value in ERROR_VALUES:
                    yield (sheet_name, cell.coordinate, str(cell.value))


def assert_no_formula_errors(path: Path) -> None:
    """Raise AssertionError if any cell contains an Excel error value."""
    errors = list(iter_error_cells(path))
    if errors:
        formatted = "\n".join(f"  {s}!{c}: {v}" for s, c, v in errors)
        raise AssertionError(
            f"Workbook {path} contains {len(errors)} formula error(s):\n{formatted}"
        )


__all__ = [
    "ERROR_VALUES",
    "recalc_workbook",
    "iter_error_cells",
    "assert_no_formula_errors",
]
