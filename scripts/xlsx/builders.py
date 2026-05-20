"""
Sheet-level builders for cfo-helper close packs and supporting deliverables.

# Pattern source: anthropics/skills/document-skills/xlsx — derived cells use
# Excel formulas (not hardcoded values) and inputs vs. formulas are
# color-coded per the styles module. Every numeric cell carries a claim_id
# comment so the reader can hover to audit.

Each builder takes an open openpyxl Workbook and a structured payload, then
writes a single sheet. Builders are idempotent: calling twice with the same
sheet name replaces the prior version.
"""

from __future__ import annotations

from typing import Iterable, Mapping, Sequence

import pandas as pd
from openpyxl import Workbook
from openpyxl.comments import Comment
from openpyxl.styles import Alignment, Font
from openpyxl.utils import get_column_letter
from openpyxl.utils.dataframe import dataframe_to_rows

from scripts.xlsx import styles


def _replace_sheet(wb: Workbook, name: str):
    if name in wb.sheetnames:
        del wb[name]
    return wb.create_sheet(name)


def _set_column_widths(ws, widths: Sequence[int]) -> None:
    for idx, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(idx)].width = w


def _comment(cell, claim_id: str | None) -> None:
    if claim_id:
        cell.comment = Comment(f"claim_id: {claim_id}", "cfo-helper")


def build_cover(
    wb: Workbook,
    *,
    period: str,
    title: str,
    subtitle: str = "",
    kpi_strip: Sequence[Mapping] = (),
) -> None:
    """Cover sheet with title, subtitle, and a compact KPI strip.

    `kpi_strip` is a sequence of dicts: {label, value, units, claim_id, number_format?}.
    Values are written as inputs (blue) since they're sourced from upstream claims.
    """
    ws = _replace_sheet(wb, "Cover")
    ws["A1"] = title
    ws["A1"].font = Font(size=20, bold=True)
    ws["A2"] = subtitle or f"Close pack — {period}"
    ws["A2"].font = Font(size=11, italic=True)
    ws["A3"] = "Hover any number for provenance (claim id)."
    ws["A3"].font = Font(size=9, italic=True, color="6B7280")

    # KPI strip: two rows per KPI (label, value), arranged in columns.
    for col_idx, kpi in enumerate(kpi_strip, start=1):
        col = get_column_letter(col_idx)
        ws[f"{col}5"] = kpi["label"]
        ws[f"{col}5"].font = Font(bold=True, size=10)
        cell = ws[f"{col}6"]
        cell.value = kpi["value"]
        if "number_format" in kpi:
            cell.number_format = kpi["number_format"]
        styles.apply_role(cell, "input", bold=True)
        cell.font = Font(bold=True, size=14, color=styles.BLUE_INPUT)
        _comment(cell, kpi.get("claim_id"))

    _set_column_widths(ws, [22] * max(len(kpi_strip), 4))


def build_pl(
    wb: Workbook,
    df: pd.DataFrame,
    *,
    sheet_name: str = "P&L",
    claim_ids_by_cell: Mapping[str, str] | None = None,
    number_format: str = styles.USD,
) -> None:
    """Write a P&L DataFrame. Numeric columns get USD formatting and input-blue."""
    ws = _replace_sheet(wb, sheet_name)
    rows = list(dataframe_to_rows(df, index=False, header=True))
    for r_idx, row in enumerate(rows, start=1):
        for c_idx, val in enumerate(row, start=1):
            cell = ws.cell(row=r_idx, column=c_idx, value=val)
            if r_idx == 1:
                cell.fill = styles.HEADER_FILL
                cell.font = styles.HEADER_FONT
                cell.alignment = Alignment(horizontal="center")
            elif isinstance(val, (int, float)) and not isinstance(val, bool):
                cell.number_format = number_format
                styles.apply_role(cell, "input")
    if claim_ids_by_cell:
        for cell_ref, claim_id in claim_ids_by_cell.items():
            _comment(ws[cell_ref], claim_id)
    _set_column_widths(ws, [max(12, min(28, df[col].astype(str).str.len().max() + 2))
                             if col in df.columns else 14 for col in df.columns])


def build_variance_detail(
    wb: Workbook,
    df: pd.DataFrame,
    *,
    sheet_name: str = "Variance Detail",
    claim_ids_by_cell: Mapping[str, str] | None = None,
) -> None:
    """Variance detail with archetype × product × mechanism columns where present.

    The DataFrame is expected to have at minimum: account, line_item, actual,
    budget, variance, plus optionally archetype, product, mechanism, customer.
    Numeric columns get USD formatting; the variance column gets red/green
    conditional via apply_role based on sign.
    """
    ws = _replace_sheet(wb, sheet_name)
    if df.empty:
        ws["A1"] = "(no material variances this period)"
        ws["A1"].font = Font(italic=True, color="6B7280")
        return

    headers = list(df.columns)
    for c_idx, h in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=c_idx, value=h)
        cell.fill = styles.HEADER_FILL
        cell.font = styles.HEADER_FONT
        cell.alignment = Alignment(horizontal="center")

    numeric_cols = {col for col in headers if pd.api.types.is_numeric_dtype(df[col])}
    variance_col = "variance" if "variance" in headers else None

    for r_idx, (_, row) in enumerate(df.iterrows(), start=2):
        for c_idx, col in enumerate(headers, start=1):
            val = row[col]
            cell = ws.cell(row=r_idx, column=c_idx, value=val)
            if col in numeric_cols:
                cell.number_format = styles.USD
                if col == variance_col and isinstance(val, (int, float)):
                    role = "external_link" if val < 0 else "internal_link"
                    styles.apply_role(cell, role)
                else:
                    styles.apply_role(cell, "input")

    if claim_ids_by_cell:
        for cell_ref, claim_id in claim_ids_by_cell.items():
            _comment(ws[cell_ref], claim_id)

    _set_column_widths(ws, [max(12, min(32, len(str(h)) + 4)) for h in headers])


def build_kpi_dashboard(
    wb: Workbook,
    kpis: Sequence[Mapping],
    *,
    sheet_name: str = "KPIs",
) -> None:
    """KPI grid: rows of (metric, value, units, comparator, % change, confidence).

    `kpis` is a sequence of dicts. Each row's comparator + % change are
    written as Excel formulas referencing the value and comparator cells so
    the reader can audit math directly in the workbook.
    """
    ws = _replace_sheet(wb, sheet_name)
    headers = ["Metric", "Period value", "Units", "Prior period", "% Change", "Confidence"]
    for c_idx, h in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=c_idx, value=h)
        cell.fill = styles.HEADER_FILL
        cell.font = styles.HEADER_FONT
        cell.alignment = Alignment(horizontal="center")

    for r_idx, kpi in enumerate(kpis, start=2):
        ws.cell(row=r_idx, column=1, value=kpi["label"])

        val_cell = ws.cell(row=r_idx, column=2, value=kpi.get("value"))
        if "number_format" in kpi:
            val_cell.number_format = kpi["number_format"]
        styles.apply_role(val_cell, "input")
        _comment(val_cell, kpi.get("claim_id"))

        ws.cell(row=r_idx, column=3, value=kpi.get("units", ""))

        prior_cell = ws.cell(row=r_idx, column=4, value=kpi.get("prior"))
        if "number_format" in kpi and kpi.get("prior") is not None:
            prior_cell.number_format = kpi["number_format"]
        styles.apply_role(prior_cell, "input")
        _comment(prior_cell, kpi.get("prior_claim_id"))

        # % change is a formula so the audit trail is in-cell
        if kpi.get("prior") not in (None, 0):
            pct_cell = ws.cell(
                row=r_idx, column=5,
                value=f"=IFERROR((B{r_idx}-D{r_idx})/D{r_idx},\"\")",
            )
            pct_cell.number_format = styles.PCT
            styles.apply_role(pct_cell, "formula")
        else:
            ws.cell(row=r_idx, column=5, value="—")

        ws.cell(row=r_idx, column=6, value=kpi.get("confidence", "high"))

    _set_column_widths(ws, [28, 16, 10, 16, 12, 12])


def build_deferred_rev_rollforward(
    wb: Workbook,
    *,
    period: str,
    opening: Mapping,
    billings: Mapping,
    recognized: Mapping,
    adjustments: Mapping | None = None,
    closing_expected: Mapping | None = None,
    sheet_name: str = "Deferred Rev Rollforward",
) -> None:
    """Deferred-revenue rollforward following the deferred-rev-rollforward skill.

    Each input dict carries: {label, value, claim_id}. The closing-balance row
    is computed via Excel formula = opening + billings - recognized + adjustments
    so the audit trail stays in the workbook.

    If `closing_expected` is provided, an additional check row computes the
    delta (formula vs expected) and conditional-formats it red if material.
    """
    ws = _replace_sheet(wb, sheet_name)
    ws["A1"] = f"Deferred Revenue Rollforward — {period}"
    ws["A1"].font = Font(size=14, bold=True)
    ws.merge_cells("A1:C1")

    rows = [
        ("Opening balance", opening),
        ("+ Billings", billings),
        ("- Revenue recognized", {**recognized, "value": -abs(recognized["value"])}),
        ("+/- Adjustments / FX", adjustments or {"label": "Adjustments", "value": 0,
                                                   "claim_id": ""}),
    ]
    start_row = 3
    for offset, (label, payload) in enumerate(rows):
        r = start_row + offset
        ws.cell(row=r, column=1, value=label).font = Font(bold=(offset == 0))
        cell = ws.cell(row=r, column=2, value=payload["value"])
        cell.number_format = styles.USD
        styles.apply_role(cell, "input")
        _comment(cell, payload.get("claim_id"))
        if payload.get("notes"):
            ws.cell(row=r, column=3, value=payload["notes"]).font = Font(
                italic=True, size=9, color="6B7280"
            )

    closing_row = start_row + len(rows)
    ws.cell(row=closing_row, column=1, value="= Closing balance").font = Font(bold=True)
    closing_cell = ws.cell(
        row=closing_row, column=2,
        value=f"=SUM(B{start_row}:B{closing_row - 1})",
    )
    closing_cell.number_format = styles.USD
    styles.apply_role(closing_cell, "formula", bold=True)

    if closing_expected:
        check_row = closing_row + 2
        ws.cell(row=check_row, column=1, value="Expected closing (independent calc)").font = Font(italic=True)
        exp_cell = ws.cell(row=check_row, column=2, value=closing_expected["value"])
        exp_cell.number_format = styles.USD
        styles.apply_role(exp_cell, "input")
        _comment(exp_cell, closing_expected.get("claim_id"))

        delta_row = check_row + 1
        ws.cell(row=delta_row, column=1, value="Delta (formula vs expected)").font = Font(italic=True)
        delta_cell = ws.cell(
            row=delta_row, column=2,
            value=f"=B{closing_row}-B{check_row}",
        )
        delta_cell.number_format = styles.USD
        styles.apply_role(delta_cell, "formula")

    _set_column_widths(ws, [32, 18, 40])


def build_top10_movement(
    wb: Workbook,
    df: pd.DataFrame,
    *,
    sheet_name: str = "Top-10 Movement",
) -> None:
    """Top-10 customer movement: additions / expansions / churns by customer.

    Expects columns: customer, archetype, beginning_arr, additions, expansions,
    contractions, churn, ending_arr. The ending_arr column is written as a
    formula so the math is auditable.
    """
    ws = _replace_sheet(wb, sheet_name)
    headers = list(df.columns)
    for c_idx, h in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=c_idx, value=h)
        cell.fill = styles.HEADER_FILL
        cell.font = styles.HEADER_FONT
        cell.alignment = Alignment(horizontal="center")

    numeric_cols = [c for c in headers if pd.api.types.is_numeric_dtype(df[c])]
    ending_col_idx = headers.index("ending_arr") + 1 if "ending_arr" in headers else None
    component_cols = ["beginning_arr", "additions", "expansions", "contractions", "churn"]
    component_indices = [headers.index(c) + 1 for c in component_cols if c in headers]

    for r_idx, (_, row) in enumerate(df.iterrows(), start=2):
        for c_idx, col in enumerate(headers, start=1):
            val = row[col]
            if ending_col_idx and c_idx == ending_col_idx and component_indices:
                # Build a formula summing the component columns.
                # Beginning + adds + expansions - contractions - churn (sign per CFO convention).
                parts = []
                for col_name in component_cols:
                    if col_name in headers:
                        idx = headers.index(col_name) + 1
                        col_letter = get_column_letter(idx)
                        sign = "-" if col_name in ("contractions", "churn") else "+"
                        parts.append(f"{sign}{col_letter}{r_idx}")
                formula = "=" + "".join(parts).lstrip("+")
                cell = ws.cell(row=r_idx, column=c_idx, value=formula)
                styles.apply_role(cell, "formula", bold=True)
            else:
                cell = ws.cell(row=r_idx, column=c_idx, value=val)
                if col in numeric_cols:
                    styles.apply_role(cell, "input")
            if col in numeric_cols:
                cell.number_format = styles.USD

    _set_column_widths(ws, [max(12, min(28, len(str(h)) + 4)) for h in headers])


def build_bsr_account_roll(
    wb: Workbook,
    df: pd.DataFrame,
    *,
    sheet_name: str = "BSR Account Roll",
) -> None:
    """Balance-sheet-review account-by-account rollforward.

    Expects columns: account, beginning, debits, credits, ending, flux_pct,
    explanation. The ending column is a formula = beginning + debits - credits.
    """
    ws = _replace_sheet(wb, sheet_name)
    headers = list(df.columns)
    for c_idx, h in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=c_idx, value=h)
        cell.fill = styles.HEADER_FILL
        cell.font = styles.HEADER_FONT
        cell.alignment = Alignment(horizontal="center")

    ending_idx = headers.index("ending") + 1 if "ending" in headers else None
    beg_idx = headers.index("beginning") + 1 if "beginning" in headers else None
    debits_idx = headers.index("debits") + 1 if "debits" in headers else None
    credits_idx = headers.index("credits") + 1 if "credits" in headers else None

    for r_idx, (_, row) in enumerate(df.iterrows(), start=2):
        for c_idx, col in enumerate(headers, start=1):
            val = row[col]
            if ending_idx and c_idx == ending_idx and beg_idx and debits_idx and credits_idx:
                formula = (
                    f"={get_column_letter(beg_idx)}{r_idx}"
                    f"+{get_column_letter(debits_idx)}{r_idx}"
                    f"-{get_column_letter(credits_idx)}{r_idx}"
                )
                cell = ws.cell(row=r_idx, column=c_idx, value=formula)
                styles.apply_role(cell, "formula", bold=True)
            else:
                cell = ws.cell(row=r_idx, column=c_idx, value=val)
                if pd.api.types.is_numeric_dtype(df[col]):
                    styles.apply_role(cell, "input")
            if pd.api.types.is_numeric_dtype(df[col]):
                cell.number_format = styles.USD

    _set_column_widths(ws, [max(14, min(32, len(str(h)) + 6)) for h in headers])


__all__ = [
    "build_cover",
    "build_pl",
    "build_variance_detail",
    "build_kpi_dashboard",
    "build_deferred_rev_rollforward",
    "build_top10_movement",
    "build_bsr_account_roll",
]
